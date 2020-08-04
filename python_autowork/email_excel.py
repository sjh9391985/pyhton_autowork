from openpyxl import load_workbook
from mail import send_email


wb = load_workbook('emails.xlsx')
ws = wb.active

for row in ws.iter_rows():
    
    name = row[0].value
    recv = row[1].value

    send_mail(name, recv)
