from openpyxl import load_workbook

wb = load_workbook('simple_result.xlsx')
ws = wb.active

area = ws['A1:B2']
for row in area:
    for cell in row:
        print(cell.value)

