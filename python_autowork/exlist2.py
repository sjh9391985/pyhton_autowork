from openpyxl import load_workbook
wb = load_workbook('simple_data.xlsx')
ws = wb.active

print(ws['A1'].value)
print(ws['A2'].value)

